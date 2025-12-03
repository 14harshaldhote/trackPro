import os
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from filelock import FileLock
from django.conf import settings
from datetime import datetime, date
import json
import uuid

class ExcelDB:
    SHEETS = {
        'TrackerDefinitions': ['tracker_id', 'name', 'time_mode', 'description', 'created_at'],
        'TaskTemplates': ['template_id', 'tracker_id', 'description', 'is_recurring', 'category', 'weight'],
        'TrackerInstances': ['instance_id', 'tracker_id', 'period_start', 'period_end', 'status'],
        'TaskInstances': ['task_instance_id', 'tracker_instance_id', 'template_id', 'description', 'status', 'date', 'notes', 'metadata'],
        'AuditLogs': ['timestamp', 'action_type', 'entity_type', 'entity_id', 'details', 'origin'],
        'Quarantine': ['quarantine_id', 'timestamp', 'original_sheet', 'original_data', 'issue_type', 'details'],
        'DayNotes': ['note_id', 'tracker_id', 'date', 'content', 'sentiment_score', 'keywords']
    }

    def __init__(self):
        self.file_path = os.path.join(settings.DATA_DIR, 'tracker.xlsx')
        self.lock_path = self.file_path + '.lock'
        self.lock = FileLock(self.lock_path)
        self._ensure_file_exists()
        
        # Caching
        self._cache = {} # sheet_name -> DataFrame
        self._indices = {} # sheet_name -> id_field -> {id_val: record}
        self._last_read_time = 0

    def _safe_save(self, wb):
        """Saves the workbook atomically to prevent corruption."""
        import shutil
        temp_path = self.file_path + '.tmp'
        try:
            wb.save(temp_path)
            os.replace(temp_path, self.file_path)
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e

    def _ensure_file_exists(self):
        """Creates the Excel file with required sheets if it doesn't exist."""
        if not os.path.exists(self.file_path):
            with self.lock:
                wb = Workbook()
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']
                
                # Create all sheets with headers
                for sheet_name, headers in self.SHEETS.items():
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(sheet_name)
                        ws.append(headers)
                
                self._safe_save(wb)
        else:
            # Ensure all sheets exist even if file exists
             with self.lock:
                wb = load_workbook(self.file_path)
                changed = False
                for sheet_name, headers in self.SHEETS.items():
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(sheet_name)
                        ws.append(headers)
                        changed = True
                if changed:
                    self._safe_save(wb)

    def _check_cache_validity(self):
        """Checks if the file has changed since last read."""
        try:
            mtime = os.path.getmtime(self.file_path)
            if mtime > self._last_read_time:
                self._cache = {}
                self._indices = {} # Clear indices
                self._last_read_time = mtime
                return False
            return True
        except OSError:
            return False

    def _read_sheet(self, sheet_name):
        """Reads a sheet into a DataFrame with caching."""
        with self.lock:
            if self._check_cache_validity() and sheet_name in self._cache:
                return self._cache[sheet_name]
                
            try:
                # engine='openpyxl' is required for .xlsx
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine='openpyxl')
                self._cache[sheet_name] = df
                return df
            except ValueError:
                # Sheet might not exist or empty
                return pd.DataFrame(columns=self.SHEETS[sheet_name])

    def _append_row(self, sheet_name, row_data):
        """Appends a row to a sheet. row_data must be a list in correct order."""
        with self.lock:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            ws.append(row_data)
            self._safe_save(wb)
            
            # Invalidate cache
            if sheet_name in self._cache:
                del self._cache[sheet_name]
            if sheet_name in self._indices:
                del self._indices[sheet_name]
                
            self._last_read_time = os.path.getmtime(self.file_path)

    def log_audit(self, action_type, entity_type, entity_id, details="", origin="system"):
        """Logs an audit entry."""
        row = [
            datetime.now().isoformat(),
            action_type,
            entity_type,
            entity_id,
            details,
            origin
        ]
        self._append_row('AuditLogs', row)

    # --- Generic CRUD Helpers ---

    def insert(self, sheet_name, data_dict):
        """
        Inserts a record into the specified sheet.
        data_dict keys must match the sheet headers.
        """
        headers = self.SHEETS[sheet_name]
        row = []
        for h in headers:
            val = data_dict.get(h)
            if isinstance(val, (dict, list)):
                val = json.dumps(val)
            elif isinstance(val, (date, datetime)):
                val = val.isoformat()
            row.append(val)
        
        self._append_row(sheet_name, row)
        
        # Log creation (except for AuditLogs itself to avoid recursion)
        if sheet_name != 'AuditLogs':
            # Try to find ID
            entity_id = 'unknown'
            for key in data_dict:
                if key.endswith('_id'):
                    entity_id = data_dict[key]
                    break
            self.log_audit('CREATE', sheet_name, entity_id, f"Created record in {sheet_name}")

    def fetch_all(self, sheet_name):
        """Returns all records from a sheet as a list of dicts."""
        df = self._read_sheet(sheet_name)
        # Replace NaN with None or empty string to avoid JSON issues
        df = df.where(pd.notnull(df), None)
        return df.to_dict('records')

    def fetch_by_id(self, sheet_name, id_field, id_value):
        """Fetches a single record by ID using O(1) index lookup."""
        # Ensure cache is valid (reloads DF if needed)
        df = self._read_sheet(sheet_name)
        if df.empty:
            return None

        # Ensure index exists
        if sheet_name not in self._indices:
            self._indices[sheet_name] = {}
        
        if id_field not in self._indices[sheet_name]:
            # Build index: {id_val: record_dict}
            # Convert DF to records first
            records = df.where(pd.notnull(df), None).to_dict('records')
            self._indices[sheet_name][id_field] = {
                str(r[id_field]): r for r in records if r.get(id_field) is not None
            }
            
        return self._indices[sheet_name][id_field].get(str(id_value))

    def fetch_filter(self, sheet_name, **filters):
        """
        Fetches records matching all filters using vectorized pandas operations.
        Example: fetch_filter('TaskInstances', status='TODO', tracker_id='123')
        """
        df = self._read_sheet(sheet_name)
        if df.empty:
            return []
            
        # Apply filters
        for key, value in filters.items():
            # Ensure column exists
            if key in df.columns:
                # Use string comparison for safety
                df = df[df[key].astype(str) == str(value)]
            else:
                return [] # Invalid filter column
        
        df = df.where(pd.notnull(df), None)
        return df.to_dict('records')

    def update_many(self, sheet_name, id_field, updates_list):
        """
        Updates multiple records in one go.
        updates_list: list of tuples (id_value, update_dict)
        """
        if not updates_list:
            return True
            
        with self.lock:
            wb = load_workbook(self.file_path)
            ws = wb[sheet_name]
            
            header = [cell.value for cell in ws[1]]
            try:
                id_idx = header.index(id_field)
            except ValueError:
                return False

            # Create a map of id -> update_dict for O(1) lookup
            # Assuming IDs are unique
            updates_map = {str(u[0]): u[1] for u in updates_list}
            
            updated_count = 0
            for row in ws.iter_rows(min_row=2):
                row_id = str(row[id_idx].value)
                if row_id in updates_map:
                    update_dict = updates_map[row_id]
                    for key, value in update_dict.items():
                        if key in header:
                            col_idx = header.index(key)
                            
                            # Serialize if needed
                            if isinstance(value, (dict, list)):
                                value = json.dumps(value)
                            elif isinstance(value, (date, datetime)):
                                value = value.isoformat()
                                
                            row[col_idx].value = value
                    updated_count += 1
                    
            if updated_count > 0:
                self._safe_save(wb)
                self._last_read_time = os.path.getmtime(self.file_path)
                if sheet_name in self._cache:
                    del self._cache[sheet_name]
                if sheet_name in self._indices:
                    del self._indices[sheet_name]
                
                if sheet_name != 'AuditLogs':
                    self.log_audit('BATCH_UPDATE', sheet_name, f"{updated_count} records", "Batch update performed")
                return True
            return False

    def update(self, sheet_name, id_field, id_value, update_dict):
        """Updates a record identified by id_field=id_value with update_dict."""
        return self.update_many(sheet_name, id_field, [(id_value, update_dict)])

    def delete(self, sheet_name, id_field, id_value):
        """
        Hard deletes a record by ID.
        Expensive operation in Excel as it requires rewriting the sheet.
        """
        with self.lock:
            wb = load_workbook(self.file_path)
            if sheet_name not in wb.sheetnames:
                return False
            
            ws = wb[sheet_name]
            header = [cell.value for cell in ws[1]]
            try:
                id_idx = header.index(id_field)
            except ValueError:
                return False

            # Find row to delete
            row_to_delete = None
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if str(row[id_idx].value) == str(id_value):
                    row_to_delete = i
                    break
            
            if row_to_delete:
                ws.delete_rows(row_to_delete)
                self._safe_save(wb)
                self._last_read_time = os.path.getmtime(self.file_path)
                if sheet_name in self._cache:
                    del self._cache[sheet_name]
                if sheet_name in self._indices:
                    del self._indices[sheet_name]
                
                if sheet_name != 'AuditLogs':
                    self.log_audit('DELETE', sheet_name, str(id_value), "Hard deleted record")
                return True
            return False
