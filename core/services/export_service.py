"""
Export Service
Handles data export in various formats (JSON, CSV, Excel)
"""
import json
import csv
from io import StringIO, BytesIO
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from core.models import TrackerDefinition, TrackerInstance, TaskInstance, TaskTemplate, UserPreferences


class ExportService:
    """Service for exporting user data in various formats"""
    
    def __init__(self, user):
        self.user = user
    
    def export_month(self, year, month, format='json', tracker_id=None):
        """
        Export data for a specific month
        
        Args:
            year: int
            month: int (1-12)
            format: 'json', 'csv', or 'xlsx'
            tracker_id: optional UUID to filter by tracker
        
        Returns:
            HttpResponse with file download
        """
        # Get date range
        from calendar import monthrange
        
        start_date = datetime(year, month, 1).date()
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day).date()
        
        # Get data
        data = self._get_month_data(start_date, end_date, tracker_id)
        
        # Export by format
        if format == 'json':
            return self._export_json(data, year, month)
        elif format == 'csv':
            return self._export_csv(data, year, month)
        elif format == 'xlsx':
            return self._export_excel(data, year, month)
        else:
            raise ValueError(f'Unsupported format: {format}')
    
    def _get_month_data(self, start_date, end_date, tracker_id=None):
        """Get aggregated data for date range"""
        tasks = TaskInstance.objects.filter(
            tracker__user=self.user,
            date__gte=start_date,
            date__lte=end_date,
            deleted_at__isnull=True
        )
        
        if tracker_id:
            tasks = tasks.filter(tracker__tracker__tracker_id=tracker_id)
        
        # Aggregate by day
        daily_data = []
        current_date = start_date
        
        while current_date <= end_date:
            day_tasks = tasks.filter(date=current_date)
            total = day_tasks.count()
            completed = day_tasks.filter(status='completed').count()
            rate = (completed / total * 100) if total > 0 else 0
            
            daily_data.append({
                'date': current_date.isoformat(),
                'day_of_week': current_date.strftime('%A'),
                'total_tasks': total,
                'completed_tasks': completed,
                'completion_rate': round(rate, 1)
            })
            
            current_date += timedelta(days=1)
        
        # Overall stats
        all_tasks = tasks.count()
        all_completed = tasks.filter(status='completed').count()
        overall_rate = (all_completed / all_tasks * 100) if all_tasks > 0 else 0
        
        return {
            'month': start_date.strftime('%B %Y'),
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'summary': {
                'total_tasks': all_tasks,
                'completed_tasks': all_completed,
                'completion_rate': round(overall_rate, 1)
            },
            'daily_data': daily_data
        }
    
    def _export_json(self, data, year, month):
        """Export as JSON"""
        response = JsonResponse(data, json_dumps_params={'indent': 2})
        filename = f'tracker_export_{year}_{month:02d}.json'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _export_csv(self, data, year, month):
        """Export as CSV"""
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['Date', 'Day of Week', 'Total Tasks', 'Completed Tasks', 'Completion Rate (%)'])
        
        # Data rows
        for day in data['daily_data']:
            writer.writerow([
                day['date'],
                day['day_of_week'],
                day['total_tasks'],
                day['completed_tasks'],
                day['completion_rate']
            ])
        
        # Summary row
        writer.writerow([])
        writer.writerow(['Summary', '', data['summary']['total_tasks'], 
                        data['summary']['completed_tasks'], 
                        data['summary']['completion_rate']])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        filename = f'tracker_export_{year}_{month:02d}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _export_excel(self, data, year, month):
        """Export as Excel using openpyxl (already in requirements.txt)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"{data['month']}"
            
            # Title
            ws['A1'] = f"Tracker Export - {data['month']}"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:E1')
            
            # Headers
            headers = ['Date', 'Day of Week', 'Total Tasks', 'Completed Tasks', 'Completion Rate (%)']
            header_fill = PatternFill(start_color='0277BD', end_color='0277BD', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row_idx, day in enumerate(data['daily_data'], 4):
                ws.cell(row=row_idx, column=1, value=day['date'])
                ws.cell(row=row_idx, column=2, value=day['day_of_week'])
                ws.cell(row=row_idx, column=3, value=day['total_tasks'])
                ws.cell(row=row_idx, column=4, value=day['completed_tasks'])
                ws.cell(row=row_idx, column=5, value=day['completion_rate'])
            
            # Summary
            summary_row = len(data['daily_data']) + 5
            ws.cell(row=summary_row, column=1, value='Summary').font = Font(bold=True)
            ws.cell(row=summary_row, column=3, value=data['summary']['total_tasks'])
            ws.cell(row=summary_row, column=4, value=data['summary']['completed_tasks'])
            ws.cell(row=summary_row, column=5, value=data['summary']['completion_rate'])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 16
            ws.column_dimensions['E'].width = 20
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'tracker_export_{year}_{month:02d}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self._export_csv(data, year, month)
